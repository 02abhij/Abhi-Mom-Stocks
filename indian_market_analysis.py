"""
indian_market_analysis.py
─────────────────────────
10-year cohort analysis: What variables drive top-gainer stocks in India?

Methodology:
  1. For each calendar year 2015–2024, identify the Top 50 price gainers
     from the universe of NSE-listed stocks (from Screener export)
  2. Tag each winner with fundamental variables (from Screener CSV)
  3. Compare winner cohort vs full universe on every variable
  4. Find which variables are systematically overrepresented in winners
  5. Track sector/theme rotation year by year

Output:
  - Excel report with full cohort analysis
  - Email with key findings summary
"""

import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime, date
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import logging
import warnings
warnings.filterwarnings('ignore')

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
log = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────
EMAIL_SENDER    = os.environ.get("EMAIL_SENDER", "02abhij@gmail.com")
EMAIL_PASSWORD  = os.environ.get("EMAIL_PASSWORD", "")
EMAIL_RECIPIENT = os.environ.get("EMAIL_RECIPIENT", "abhi@ajassoc.in")
SMTP_HOST       = "smtp.gmail.com"
SMTP_PORT       = 587

YEARS           = list(range(2015, 2025))   # 2015 to 2024
TOP_N_WINNERS   = 50                         # Top gainers per year
MIN_PRICE       = 10                         # Filter sub-₹10 stocks
MIN_MCAP        = 100                        # ₹ Cr minimum market cap at start of year
BATCH_SIZE      = 100                        # yfinance download batch size

FUNDAMENTALS_CSV = "screener_fundamentals.csv"


# ── Step 1: Load Screener fundamentals ───────────────────────────────────────

def load_fundamentals() -> pd.DataFrame:
    df = pd.read_csv(FUNDAMENTALS_CSV)
    df = df[df['NSE Code'].notna() & (df['NSE Code'].astype(str).str.strip() != '')]
    df['ticker'] = df['NSE Code'].astype(str).str.strip() + '.NS'
    log.info(f"Loaded {len(df)} stocks from Screener")
    return df


# ── Step 2: Download annual price data ───────────────────────────────────────

def download_annual_returns(tickers: list[str]) -> pd.DataFrame:
    """
    For each ticker, download Jan 1 and Dec 31 price for each year 2015–2024.
    Returns DataFrame: ticker × year → annual_return
    """
    log.info(f"Downloading price history for {len(tickers)} tickers...")

    # Download full history in batches
    all_prices = {}
    batches = [tickers[i:i+BATCH_SIZE] for i in range(0, len(tickers), BATCH_SIZE)]

    for i, batch in enumerate(batches, 1):
        log.info(f"  Price batch {i}/{len(batches)}...")
        try:
            raw = yf.download(
                batch,
                start="2014-12-15",
                end="2025-01-15",
                auto_adjust=True,
                progress=False,
                threads=True,
            )
            if len(batch) == 1:
                t = batch[0]
                if not raw.empty:
                    all_prices[t] = raw['Close']
            else:
                for t in batch:
                    try:
                        s = raw['Close'][t].dropna()
                        if len(s) > 100:
                            all_prices[t] = s
                    except Exception:
                        pass
        except Exception as e:
            log.warning(f"Batch {i} failed: {e}")

    log.info(f"Got price history for {len(all_prices)} tickers")

    # Compute annual returns for each year
    records = []
    for ticker, prices in all_prices.items():
        row = {'ticker': ticker}
        valid = True
        for year in YEARS:
            try:
                # Get first and last trading day of the year
                year_prices = prices[
                    (prices.index >= f"{year}-01-01") &
                    (prices.index <= f"{year}-12-31")
                ]
                if len(year_prices) < 50:  # need reasonable trading days
                    row[f'return_{year}'] = np.nan
                    row[f'start_price_{year}'] = np.nan
                    continue
                start_price = float(year_prices.iloc[0])
                end_price = float(year_prices.iloc[-1])
                if start_price < MIN_PRICE:
                    row[f'return_{year}'] = np.nan
                else:
                    row[f'return_{year}'] = (end_price - start_price) / start_price * 100
                row[f'start_price_{year}'] = start_price
            except Exception:
                row[f'return_{year}'] = np.nan
                row[f'start_price_{year}'] = np.nan
        records.append(row)

    return pd.DataFrame(records)


# ── Step 3: Identify top gainers per year ────────────────────────────────────

def get_winners_per_year(returns_df: pd.DataFrame) -> dict[int, pd.DataFrame]:
    """Returns dict: year → DataFrame of top 50 winners"""
    winners = {}
    for year in YEARS:
        col = f'return_{year}'
        yr = returns_df[['ticker', col]].dropna(subset=[col])
        yr = yr[yr[col] > 0]  # only positive returners
        top = yr.nlargest(TOP_N_WINNERS, col).copy()
        top['rank'] = range(1, len(top) + 1)
        top['year'] = year
        winners[year] = top
        log.info(f"{year}: Top gainer = {top.iloc[0]['ticker']} "
                 f"({top.iloc[0][col]:.0f}%), universe = {len(yr)} stocks")
    return winners


# ── Step 4: Tag winners with fundamentals ────────────────────────────────────

def tag_with_fundamentals(winners: dict, fundamentals: pd.DataFrame) -> pd.DataFrame:
    """Join winner cohorts with Screener fundamentals"""
    all_rows = []
    for year, df in winners.items():
        merged = df.merge(fundamentals, on='ticker', how='left')
        all_rows.append(merged)
    return pd.concat(all_rows, ignore_index=True)


# ── Step 5: Variable analysis ─────────────────────────────────────────────────

NUMERIC_VARS = [
    'Price to Earning', 'Return on equity', 'Debt to equity',
    'Sales growth 5Years', 'Profit growth 5Years', 'Promoter holding',
    'Change in promoter holding 3Years', 'Pledged percentage',
    'Dividend yield', 'Price to book value', 'Market Capitalization'
]

def analyze_variables(tagged_winners: pd.DataFrame,
                      fundamentals: pd.DataFrame) -> pd.DataFrame:
    """
    For each fundamental variable, compare:
      - Median value in winner cohort
      - Median value in full universe
      - Winner premium/discount
    """
    results = []
    for var in NUMERIC_VARS:
        if var not in tagged_winners.columns:
            continue
        winner_vals = tagged_winners[var].dropna()
        universe_vals = fundamentals[var].dropna() if var in fundamentals.columns else pd.Series()
        if len(winner_vals) < 10:
            continue
        winner_median = winner_vals.median()
        universe_median = universe_vals.median() if len(universe_vals) > 0 else np.nan
        if universe_median and universe_median != 0:
            premium = (winner_median - universe_median) / abs(universe_median) * 100
        else:
            premium = np.nan
        results.append({
            'Variable': var,
            'Winner Median': round(winner_median, 2),
            'Universe Median': round(universe_median, 2) if pd.notna(universe_median) else 'N/A',
            'Winner Premium %': round(premium, 1) if pd.notna(premium) else 'N/A',
            'Winner Count': len(winner_vals),
        })
    result_df = pd.DataFrame(results)
    result_df['Winner Premium %'] = pd.to_numeric(result_df['Winner Premium %'], errors='coerce')
    return result_df.sort_values('Winner Premium %', ascending=False)


def sector_rotation(tagged_winners: pd.DataFrame) -> pd.DataFrame:
    """Which sectors dominated each year?"""
    pivot = tagged_winners.groupby(['year', 'Industry Group']).size().reset_index(name='count')
    # Get top 3 sectors per year
    top3 = pivot.sort_values(['year','count'], ascending=[True,False])
    top3 = top3.groupby('year').head(3).reset_index(drop=True)
    return top3


def yearly_stats(tagged_winners: pd.DataFrame) -> pd.DataFrame:
    """Summary stats per year"""
    rows = []
    for year in YEARS:
        yr = tagged_winners[tagged_winners['year'] == year]
        col = f'return_{year}'
        if col not in yr.columns:
            continue
        rows.append({
            'Year': year,
            'Median Return %': round(yr[col].median(), 1),
            'Min Return %': round(yr[col].min(), 1),
            'Max Return %': round(yr[col].max(), 1),
            'Avg PE': round(yr['Price to Earning'].median(), 1),
            'Avg ROE %': round(yr['Return on equity'].median(), 1),
            'Avg D/E': round(yr['Debt to equity'].median(), 2),
            'Avg Promoter %': round(yr['Promoter holding'].median(), 1),
            'Top Sector': yr['Industry Group'].value_counts().index[0]
                          if yr['Industry Group'].notna().any() else 'N/A',
        })
    return pd.DataFrame(rows)


# ── Step 6: Build Excel report ───────────────────────────────────────────────

def build_excel(tagged_winners, fundamentals, var_analysis, sector_rot, yr_stats):
    wb = Workbook()

    BLUE  = "1e40af"
    LBLUE = "dbeafe"
    GOLD  = "f59e0b"
    WHITE = "ffffff"
    DGREY = "374151"
    LGREY = "f9fafb"

    header_font  = Font(name='Arial', bold=True, color=WHITE, size=11)
    header_fill  = PatternFill('solid', start_color=BLUE)
    subhdr_font  = Font(name='Arial', bold=True, color=DGREY, size=10)
    subhdr_fill  = PatternFill('solid', start_color=LBLUE)
    normal_font  = Font(name='Arial', size=10)
    bold_font    = Font(name='Arial', bold=True, size=10)
    alt_fill     = PatternFill('solid', start_color=LGREY)
    center       = Alignment(horizontal='center', vertical='center')
    left         = Alignment(horizontal='left', vertical='center')

    def header_row(ws, row, cols, values):
        for col, val in zip(cols, values):
            c = ws.cell(row=row, column=col, value=val)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 1: Summary ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "📊 Summary by Year"
    ws1.row_dimensions[1].height = 30
    header_row(ws1, 1, range(1, len(yr_stats.columns)+1), yr_stats.columns.tolist())
    for i, row in yr_stats.iterrows():
        r = i + 2
        fill = alt_fill if i % 2 == 0 else None
        for j, val in enumerate(row.values, 1):
            c = ws1.cell(row=r, column=j, value=val)
            c.font = normal_font
            c.alignment = center
            if fill:
                c.fill = fill
    set_col_widths(ws1, [8, 14, 12, 12, 10, 10, 10, 16, 22])
    ws1.freeze_panes = 'A2'

    # ── Sheet 2: Variable Analysis ────────────────────────────────
    ws2 = wb.create_sheet("🔬 Variable Analysis")
    ws2.row_dimensions[1].height = 30

    # Title
    ws2['A1'] = "What separates winners from the market? (All years combined)"
    ws2['A1'].font = Font(name='Arial', bold=True, size=13, color=BLUE)
    ws2.merge_cells('A1:E1')

    header_row(ws2, 2, range(1, len(var_analysis.columns)+1), var_analysis.columns.tolist())
    for i, row in var_analysis.iterrows():
        r = i + 3
        fill = alt_fill if i % 2 == 0 else None
        for j, val in enumerate(row.values, 1):
            c = ws2.cell(row=r, column=j, value=val)
            c.font = normal_font
            c.alignment = center
            if fill:
                c.fill = fill
            # Colour premium column
            if j == 4 and isinstance(val, (int, float)):
                if val > 20:
                    c.font = Font(name='Arial', bold=True, size=10, color="16a34a")
                elif val < -20:
                    c.font = Font(name='Arial', bold=True, size=10, color="dc2626")
    set_col_widths(ws2, [30, 16, 18, 16, 14])
    ws2.freeze_panes = 'A3'

    # ── Sheet 3: Sector Rotation ──────────────────────────────────
    ws3 = wb.create_sheet("🔄 Sector Rotation")
    ws3['A1'] = "Top 3 sectors in winner cohort each year"
    ws3['A1'].font = Font(name='Arial', bold=True, size=13, color=BLUE)
    ws3.merge_cells('A1:C1')
    header_row(ws3, 2, [1,2,3], ['Year', 'Sector', 'Winners Count'])
    for i, row in sector_rot.iterrows():
        r = i + 3
        fill = alt_fill if i % 2 == 0 else None
        for j, val in enumerate(row.values, 1):
            c = ws3.cell(row=r, column=j, value=val)
            c.font = normal_font
            c.alignment = center
            if fill:
                c.fill = fill
    set_col_widths(ws3, [8, 35, 16])
    ws3.freeze_panes = 'A3'

    # ── Sheet 4: Full Winner Database ─────────────────────────────
    ws4 = wb.create_sheet("📋 All Winners")
    display_cols = ['year', 'rank', 'ticker', 'Name', 'Industry Group',
                    f'return_{YEARS[0]}',  # placeholder, will fix below
                    'Price to Earning', 'Return on equity', 'Debt to equity',
                    'Sales growth 5Years', 'Profit growth 5Years',
                    'Promoter holding', 'Change in promoter holding 3Years',
                    'Pledged percentage', 'Market Capitalization']

    # Build dynamic return column per row
    out_cols = ['Year', 'Rank', 'Ticker', 'Name', 'Sector',
                'Annual Return %', 'PE', 'ROE %', 'D/E',
                'Sales CAGR 5Y %', 'PAT CAGR 5Y %',
                'Promoter %', 'Promoter Chg 3Y %',
                'Pledged %', 'Mkt Cap (Cr)']
    header_row(ws4, 1, range(1, len(out_cols)+1), out_cols)
    ws4.row_dimensions[1].height = 28

    r = 2
    for _, row in tagged_winners.sort_values(['year','rank']).iterrows():
        year = int(row['year'])
        ret_col = f'return_{year}'
        ret_val = row.get(ret_col, np.nan)
        vals = [
            year, row.get('rank'), row.get('ticker','').replace('.NS',''),
            row.get('Name'), row.get('Industry Group'),
            round(ret_val, 1) if pd.notna(ret_val) else '',
            row.get('Price to Earning'), row.get('Return on equity'),
            row.get('Debt to equity'), row.get('Sales growth 5Years'),
            row.get('Profit growth 5Years'), row.get('Promoter holding'),
            row.get('Change in promoter holding 3Years'),
            row.get('Pledged percentage'), row.get('Market Capitalization'),
        ]
        fill = alt_fill if (r % 2 == 0) else None
        for j, val in enumerate(vals, 1):
            c = ws4.cell(row=r, column=j, value=val if pd.notna(val) else '')
            c.font = normal_font
            c.alignment = center
            if fill:
                c.fill = fill
            # Colour return column
            if j == 6 and isinstance(val, (int, float)):
                c.font = Font(name='Arial', bold=True, size=10,
                              color="16a34a" if val > 0 else "dc2626")
        r += 1

    set_col_widths(ws4, [8,7,14,28,28,14,8,9,8,14,14,12,16,11,14])
    ws4.freeze_panes = 'A2'

    path = "indian_market_analysis.xlsx"
    wb.save(path)
    log.info(f"Excel saved: {path}")
    return path


# ── Step 7: Email ─────────────────────────────────────────────────────────────

def send_email(excel_path: str, yr_stats: pd.DataFrame, var_analysis: pd.DataFrame):
    from email.mime.base import MIMEBase
    from email import encoders

    # Build summary text
    best_years = yr_stats.nlargest(3, 'Median Return %')[['Year','Median Return %','Top Sector']]
    top_vars = var_analysis.head(4)['Variable'].tolist() if len(var_analysis) > 0 else []

    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:700px;margin:0 auto;">
      <div style="background:linear-gradient(135deg,#1e40af,#3b82f6);padding:28px;color:white;border-radius:12px 12px 0 0;">
        <div style="font-size:22px;font-weight:800;">🇮🇳 Indian Market Return Analysis</div>
        <div style="font-size:14px;opacity:0.85;margin-top:4px;">10-year cohort study · 2015–2024 · Top 50 gainers per year</div>
      </div>
      <div style="padding:24px;background:white;border:1px solid #e5e7eb;">
        <p style="font-size:15px;color:#374151;">
          Analysis complete. Full Excel report attached with 4 sheets:
          <b>Summary by Year, Variable Analysis, Sector Rotation, All Winners database.</b>
        </p>

        <h3 style="color:#1e40af;">📅 Best years for momentum</h3>
        <table style="border-collapse:collapse;width:100%;font-size:13px;">
          <tr style="background:#dbeafe;">
            <th style="padding:8px;text-align:left;">Year</th>
            <th style="padding:8px;text-align:center;">Median Top-50 Return</th>
            <th style="padding:8px;text-align:left;">Dominant Sector</th>
          </tr>
          {''.join(f"<tr><td style='padding:8px;'>{int(r.Year)}</td><td style='padding:8px;text-align:center;font-weight:bold;color:#16a34a;'>{r['Median Return %']}%</td><td style='padding:8px;'>{r['Top Sector']}</td></tr>" for _,r in best_years.iterrows())}
        </table>

        <h3 style="color:#1e40af;margin-top:20px;">🔬 Variables overrepresented in winners</h3>
        <p style="color:#374151;font-size:13px;">
          These fundamentals appeared significantly more often (or at higher levels)
          in the top-50 gainer cohort vs the broader universe:
        </p>
        <ul style="color:#374151;font-size:13px;">
          {''.join(f"<li><b>{v}</b></li>" for v in top_vars)}
        </ul>

        <p style="font-size:12px;color:#9ca3af;margin-top:20px;">
          See the <b>Variable Analysis</b> sheet for full winner vs universe comparison.<br>
          See <b>Sector Rotation</b> for year-by-year theme breakdown.<br>
          Not investment advice.
        </p>
      </div>
    </div>
    """

    msg = MIMEMultipart()
    msg['Subject'] = "🇮🇳 Indian Market Return Analysis · 2015–2024"
    msg['From'] = EMAIL_SENDER
    msg['To'] = EMAIL_RECIPIENT
    msg.attach(MIMEText(html, 'html'))

    # Attach Excel
    with open(excel_path, 'rb') as f:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="indian_market_analysis.xlsx"')
        msg.attach(part)

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, EMAIL_RECIPIENT, msg.as_string())
        log.info("✅ Email sent with Excel attachment")
    except Exception as e:
        log.error(f"Email failed: {e}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    log.info("=" * 60)
    log.info("INDIAN MARKET RETURN ANALYSIS — 2015 to 2024")
    log.info("=" * 60)

    # 1. Load fundamentals
    fundamentals = load_fundamentals()

    # 2. Download price history
    tickers = fundamentals['ticker'].tolist()
    returns_df = download_annual_returns(tickers)

    # 3. Get winners per year
    winners = get_winners_per_year(returns_df)

    # 4. Tag with fundamentals
    tagged = tag_with_fundamentals(winners, fundamentals)
    log.info(f"Total winner-year observations: {len(tagged)}")

    # 5. Analyse
    var_analysis = analyze_variables(tagged, fundamentals)
    sector_rot   = sector_rotation(tagged)
    yr_stats     = yearly_stats(tagged)

    log.info("\n" + "="*50)
    log.info("VARIABLE ANALYSIS — Winners vs Universe")
    log.info("="*50)
    log.info("\n" + var_analysis.to_string(index=False))

    log.info("\nSECTOR ROTATION")
    log.info("\n" + sector_rot.to_string(index=False))

    # 6. Build Excel
    excel_path = build_excel(tagged, fundamentals, var_analysis, sector_rot, yr_stats)

    # 7. Email
    send_email(excel_path, yr_stats, var_analysis)

    log.info("DONE.")


if __name__ == "__main__":
    main()
